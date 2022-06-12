






## Algorithme KNN


from math import sqrt
from math import floor


def k(dataset):
    k = floor(sqrt(len(dataset)))
    if k%2== 0: k += 1
    return k

def distance(a,dataset):
    t = []
    for k in range(len(dataset)):
        res = 0
        for l in range(len(a)):
            res += (a[l]-dataset[k][l])**2
        t.append([k,sqrt(res)])
    return t


def distance_point(a,b):
    res = 0
    for l in range(len(a)):
        res += importances[l]*(a[l]-b[l])**2
    return (sqrt(res))

def tri_insertion(liste,x):
    liste.append(x)
    i = len(liste)-1
    while i > 0 and x[1] < liste[i-1][1] :
        liste[i] = liste[i-1]
        liste[i-1] = x
        i = i-1
    return liste

def plus_proches(a,dataset,k):
    plus_proches = []
    for i in range (k):
        tri_insertion(plus_proches,[i,distance_point(a,dataset[i])])
    for i in range (k,len(dataset)):
        tri_insertion(plus_proches,[i,distance_point(a,dataset[i])])
        plus_proches.remove(plus_proches[-1])
    return plus_proches


def classe_voisins (a,dataset,k):
    c = [dataset[i][-1] for [i,b] in plus_proches(a,dataset,k)]
    return c

def classe_majoritaire(c):
    t = []
    for k in range (len(c)):
        if c[k] not in t : t.append(c[k])
    m = [0 for i in range(len(t))]
    for k in range (len(c)):
        a = t.index(c[k])
        m[a] += 1
    return t[m.index(max(m))]


def classe_majoritaire_distance(a,dataset,k):
    t = []
    voisins = plus_proches(a,dataset,k)
    classe_voisins = []
    distance_voisins = []
    for k in range (len(voisins)):
        Id,d = voisins[k]
        classe_voisins.append(dataset[Id][-1])
        distance_voisins.append(d)

    for k in range (len(classe_voisins)):
        if classe_voisins[k] not in t : t.append(classe_voisins[k])

    m = [0 for i in range(len(t))]
    for k in range (len(distance_voisins)):
        b = t.index(classe_voisins[k])
        d = distance_voisins[k]
        if d != 0 :
            m[b] += ( 10/((d)**3) )
        else : m[b]+= 10
    return [ t[m.index(max(m))],m ]

def percentage_prediction_distance (a,dataset,k):
    [prediction,percentage] = classe_majoritaire_distance(a,dataset,k)
    if prediction == 'B' : autre = 'M'
    else : autre = 'B'
    if len(percentage) == 1 : percentage.append(0)
    sum_percentage = percentage[0]+percentage[1]
    percentage[0] = percentage[0]/sum_percentage
    percentage[1] = percentage[1]/sum_percentage
    return [ (prediction,max(percentage)),(autre,min(percentage))]

def prediction (a,dataset,k):
    p = (classe_majoritaire(classe_voisins(a,dataset,k)))
    return p

def prediction_distance(a,dataset,k):
    p = (classe_majoritaire_distance(a,dataset,k)[0])
    return p


## Amelioration dataset

# def viable (a,dataset,k):
#     b = plus_proches(a,dataset,k)
#     c = classe_voisins(a,dataset,k)
#     m = c[0]
#     for k in range (len(c)):
#         if c[k] != m : return False
#     for k in range (len(b)):
#         if distance(a,dataset)[b[k]] > 1.5 : return False
#     return True

def tri_param (dataset,l):
    t = []
    for k in range (len(dataset)):
        t.append([dataset[k][i] for i in l])
    return t


param = [0,2,3,5,6,7,12,13,15,17,20,22,23,25,26,27,-1]

# on remarque des resultats égaux avec moins de paramètre ( 17 au lieu de 30 ) et un gain de temps et quasi identiques avec 5 paramètres, bien choisis


def moyenne_dataset(dataset):
    M = [0 for i in range (len(dataset[0])-1)]
    for k in range (len(dataset)):
        for l in range (len(dataset[k])-1):
            M[l] += dataset[k][l]
    for k in range(len(M)):
        M[k] = M[k]/(len(dataset))
    return M

def ecart_type (dataset):
    ecart = []
    m = moyenne_dataset(dataset)
    for l in range (len(dataset[0])-1):
        e = 0
        for k in range (len(dataset)):
            a = (m[l]-dataset[k][l])
            e += a*a
        ecart.append(sqrt(e/(len(dataset))))
    return ecart

def standardization (dataset):
    data = []
    m = moyenne_dataset(dataset)
    e = ecart_type(dataset)
    for k in range (len(dataset)):
        d = []
        for l in range (len(dataset[0])-1) :
           d.append((dataset[k][l] - m[l])/e[l])
        d.append(dataset[k][-1])
        data.append(d)
    return data




## Visualisation importance paramètres

import matplotlib.pyplot as plt

def trace_parametre(a,dataset):
    A = []
    B = []
    D = []
    C = []
    for k in range (len(dataset)):
        if dataset[k][-1] == 'M' :
            A.append(k)
            B.append(dataset[k][a])
        else :
            C.append(k)
            D.append(dataset[k][a])
    plt.scatter(A,B,c = 'red',label='M')
    plt.scatter(C,D,c = 'blue',label='B')
    plt.title('Nuage de points avec Matplotlib')
    plt.xlabel('x')
    plt.ylabel('y')
    plt.show()

def moyenne (dataset):
    M = [0 for i in range (len(dataset[k])-1)]
    B = [0 for i in range (len(dataset[k])-1)]
    m = 0
    b = 0
    for k in range (len(dataset)):
        if dataset[k][-1]=='M':
            for l in range (len(dataset[k])-1):
                M[l] += dataset[k][l]
                m += 1
        else :
            for l in range (len(dataset[k])-1):
                B[l] += dataset[k][l]
                b += 1
    for k in range(len(M)):
        M[k] = M[k]/m
    for i in range (len(B)):
        B[k] = B[k]/b
    return (M,B)

def rapport_moyenne(dataset):
    t = []
    A,B = moyenne(dataset)
    for k in range (len(A)):
        t.append(A[k]/B[k])
    return t


def trace_moyenne(dataset):
    a = [ i for i in range (30)]
    B,C = moyenne(dataset)
    plt.scatter(a,B,c = 'red',label='M')
    plt.scatter(a,C,c = 'blue',label='B')
    plt.title('Nuage de points avec Matplotlib')
    plt.xlabel('x')
    plt.ylabel('y')
    plt.show()


def trace_test(l):
    b = [l[k] for k in range (len(l)-1)]
    a = [ i for i in range (29)]
    plt.scatter(a,b,c = 'blue',label='Rapport')
    plt.title('Nuage de points avec Matplotlib')
    plt.xlabel('x')
    plt.ylabel('y')
    plt.show()







## Tests et optimisation

# Avec une dataset de 100 éléments : pourcentage d'erreur = 12.0%  ( test sur 100 éléments )
#                     150                                 ~  9.0%
#                     200                                 ~  9.0%
#                     250                                 ~  9.0%
#                     300                                 ~ ..%
#                     350                                 ~ ..%
#                     400                                 ~ ..%
#                     450                                 ~ ..%
#                     500                                 ~ ..%


# en améliorant le dataset : - enlever les données abbérantes et normaliser les donnés restantes, on obtient un résultat de 7,14% d'erreur avec une base de données 250 données ce qui montre l'importance de la normalisation. ( ici k = 8 )
# en triant les paramètres , avec k = 6 on passe à 8,16%, la perte de données est trop couteuse en précison.
# Pourtant avec k = 4 on repasse à 7,14%.



# Avec la distance coefficiantant l'importance des données sur le query on obtient un
# un résultat de 6,12% d'erreur.


elt = [110,150,200,250,300,350,400,450,500]
#err = [12,9,9,..,..,..,..,..,..]

def trace(a,b):
    plt.scatter(a,b,c = 'blue')
    plt.title('Nuage de points avec Matplotlib')
    plt.xlabel("Nombre d'elements")
    plt.ylabel("Pourcentage d'erreurs")
    plt.show()



test = [17.05,19.08,113.4,895,0.1141,0.1572,0.191,0.109,0.2131,0.06325,0.2959,0.679,2.153,31.98,0.005532,0.02008,0.03055,0.01384,0.01177,0.002336,19.59,24.89,133.5,1189,0.1703,0.3934,0.5018,0.2543,0.3109,0.09061]
# test exp est M --> prediction = M

test2 = [11.32,27.08,71.76,395.7,0.06883,0.03813,0.01633,0.003125,0.1869,0.05628,0.121,0.8927,1.059,8.605,0.003653,0.01647,0.01633,0.003125,0.01537,0.002052,12.08,33.75,79.82,452.3,0.09203,0.1432,0.1089,0.02083,0.2849,0.07087]
# test2 exp est B --> predction = B


def reponse_test (a):
  return [a[k][-1] for k in range(len(a))]

def datatest(dataset):
    datatest = []
    for k in range (len(dataset)):
        m = []
        for l in range (len(dataset[k])-1):
            m.append(dataset[k][l])
        datatest.append(m)
    return datatest

def resultat_test(test,dataset,nbr_voisins):
    t = []
    for k in range (len(test)):
        t.append(prediction(datatest(test)[k],dataset,nbr_voisins))
    c=0
    for k in range(len(t)):
        if t[k] != test[k][-1] : c += 1
    return (c,(c/len(t))*100)

def resultat_test_distance(test,dataset,nbr_voisins):
    t = []
    for k in range (len(test)):
        t.append(prediction_distance(datatest(test)[k],dataset,nbr_voisins))
    c=0
    for k in range(len(t)):
        if t[k] != test[k][-1] :
            c += 1
    return (c,(c/len(t))*100)

def moyenne (liste):
    l = len(liste)
    res = 0
    for k in range (l):
        res += liste[k]
    return res/l


def k_fold_verif (k,testdata,resultat,nbr_voisins):
    l = len(testdata)
    h = l//k
    list_result = []
    for i in range (k):
        data_test = testdata[i*h:(i+1)*h]
        data_donnée = testdata[:i*h] + testdata[(i+1)*h:]
        a,b = resultat(data_test,data_donnée,nbr_voisins)
        list_result.append(b)
    mean_list = moyenne(list_result)
    return mean_list


def perf_nombre_voisins (donnée,resultat):
    performances = []
    p = floor(sqrt(len(donnée)))+1
    fold = len(donnée)//100
    for k in range (1,p):
        res = k_fold_verif (fold,donnée,resultat,k)
        performances.append(res)
    return performances


def automatisation_nbr_voisin (dataset,n):
    l = len(dataset)
    a = floor(0.70*l)
    fold = a//100
    p = floor(sqrt(a))+1
    resultat = [ 0 for i in range (1,p)]
    for i in range (n):
        data = shuffle(dataset)
        data_donnée = data[:a]
        performances = perf_nombre_voisins(data_donnée,resultat_test_distance)
        for u in range(len(performances)):
            resultat[u] += performances[u]
    nb_voisins = resultat.index(min(resultat))+1
    perf_nb_voisins = resultat[nb_voisins -1]/n
    return (nb_voisins,perf_nb_voisins)

def eval_perf (dataset,nb_voisins,n):
    l = len(dataset)
    a = floor(0.70*l)
    fold = a//100
    res = 0
    res2 = 0
    for i in range (n):
        data = shuffle(dataset)
        data_donnée = data[:a]
        data_test = data[a:]
        res  += k_fold_verif(fold,data_donnée,resultat_test,nb_voisins)
        #res += resultat_test(data_test,data_donnée,nb_voisins)[1]
        res2 += resultat_test_distance(data_test,data_donnée,nb_voisins)[1]
    return res/n, res2/n


#                       nb_voisins      % erreur moyenne
# automatisation_wis =      6               3.81 %
# automatisation_mass =    24               21.85%

# avec 6 voisins, pour wis : moyenne a 3,81% d'erreur pour distance et 4,11% pour normal
# avec 24 voisins, pour wis : moyenne a 21,85% d'erreur pour distance et 20,48% pour normal


nb_wis = 6
nb_mass = 24

## Sortie et ROC

def sortie_algo(test,dataset,nbr_voisins):
    t = []
    for k in range (len(test)):
        t.append(prediction(datatest(test)[k],dataset,nbr_voisins))
    return t


def sortie_algo_distance(test,dataset,nbr_voisins):
    t = []
    for k in range (len(test)):
        t.append(prediction_distance(datatest(test)[k],dataset,nbr_voisins))
    return t


def ROC (test,dataset,sortie,nbr_voisins):
    L = test
    l = len(L)
    a = sortie(test,dataset,nbr_voisins)
    b = reponse_test(L)
    x = []
    y = []
    precision = []
    rappel = []
    FN = 0
    TP = 0
    TN = 0
    FP = 0
    for k in range (l):
        if b[k]=='B' :
            if a[k]=='B' : TP += 1
            else : FN += 1
            y.append(TP)
            x.append(FP)
        else :
            if a[k]=='M' : TN += 1
            else : FP += 1
            x.append(FP)
            y.append(TP)

    specificite__1 = [(x[k]/(FP+TN))  for k in range (len(x))]
    specificite_1 = [ specificite__1[k]/specificite__1[-1] for k in range (len(specificite__1))]
    sensitivite1 = [((y[k]/(TP+FN))/(TP/TP+FN)) for k in range (len(y))]
    sensitivite = [sensitivite1[k]/sensitivite1[-1] for k in range (len(sensitivite1))]
    precision = [y[k]/(TP+FP) for k in range (len(y))]
    rappel = [y[k]/(TP+FN) for k in range (len(y))]

    for k in range (len(x)):
        x[k]=x[k]/(FP+TN)
    for k in range (len(y)):
        y[k]=y[k]/(FN+TP)
    for k in range (len(y)):
        y[k]=y[k]/y[-1]
    x.append(1)
    y.append(1)

    PIPP = TP/(TP+FP)   # positifs_in_predits_positifs ( précision )
    PPIP = TP/(TP+FN)   # predits_positifs_in_positifs ( rappel )
    MH = 2*(PIPP*PPIP)/(PIPP+PPIP)  #moyenne_harmonique
    return [PIPP,PPIP,MH,(specificite_1,sensitivite),(precision,rappel),(x,y)]

# Taux faux positifs = FP/(FP+TN)   ( 1 - spécificité )
# Taux vrais positifs = TP/(TP+FN)  ( sensitivité )

# On obtient :
# PIPP = 0.9629629629629629
# PPIP = 0.9285714285714286
#  MH  = 0.9454545454545454

def graphe_ROC(test,dataset,sortie,nbr_voisins):
    x,y = ROC(test,dataset,sortie,nbr_voisins)[-1]
    plt.plot(x,y,c = 'red',label='ROC')
    plt.title('Courbe de ROC')
    plt.xlabel('taux faux positifs')
    plt.ylabel('taux vrais positifs')
    plt.show()

def graphe_ROC_specificite_sensitivite (test,dataset,sortie,nbr_voisins): # pas dingue
    specificite_1,sensitivite = ROC(test,dataset,sortie,nbr_voisins)[-3]
    plt.plot(specificite_1,sensitivite,c = 'red')
    plt.plot([0,1],[0,1],c='b')
    plt.title('sensitivite = f(1 - specificite)')
    plt.xlabel('1 - specificite')
    plt.ylabel('sensitivite')
    plt.show()

def graphe_ROC_precision_rappel (test,dataset,sortie,nbr_voisins):
    precision,rappel = ROC(test,dataset,sortie,nbr_voisins)[-2]
    plt.plot(rappel,precision,c = 'red',label='ROC')
    plt.title('precison = f(rappel)')
    plt.xlabel('rappel')
    plt.ylabel('precision')
    plt.show()

from random import *

def aire_ROC(test,dataset,sortie,n,nbr_voisins):   # on utilise ici la méthode de monte carlo
    x,y = ROC(test,dataset,sortie,nbr_voisins)[-1]
    c = 0
    for k in range (n):
        x0 = random()
        y0 = random()
        k = 0
        xk = 0
        while k < len(x) and xk < x0 :
            k = k+1
            xk = x[k]
        if y0 > y[k] : c += 1
    return 1-c/n

# on obtient une aire de 0.982099 pour n = 10^6
# cette aire très importante est due à la taille réduite du dataset et donc au petit nombre d'erreur, ainsi la position des erreurs influe enormement,une erreur des les premieres données donne une aire plus petite et inversement.
# Il faut plus de données pour une aire plus crédible.

# On a maintenant beaucoup plus de données et la nouvelle aire est credible ,
# On obtient pour n = 10^6 : aire = 0.989076

## Bases de données



import random

import csv

def lecture_fichier():
    with open ('dataset_convert.csv','r') as base_donnee :
        csv_reader = csv.reader(base_donnee)
        dataset = []
        for line in (base_donnee):
            dataset.append(line.split(","))
        for k in range (len(dataset)):
            for l in range (len(dataset[k])):
                dataset[k][l] = float(dataset[k][l])
        return dataset



def mise_en_forme(dataset):
    data = [dataset[k][2:]+[dataset[k][1],dataset[k][0]]for k in range (len(dataset))]
    return data

def mise_en_forme_temp(dataset):
    data = [dataset[k][2:]+['M' if dataset[k][1]==1 else 'B']for k in range (len(dataset))]
    return data

def shuffle (liste_originale):
    liste = liste_originale[::]
    l = len(liste)
    liste_shuffle = []
    for k in range (l):
        n = (l-1-k)
        r = random.randint(0,n)
        liste_shuffle.append(liste[r])
        liste.remove(liste[r])
    return liste_shuffle

Base_de_données_wisconsin_originale = mise_en_forme_temp(lecture_fichier())
Base_de_données_wisconsin = shuffle(Base_de_données_wisconsin_originale)

stand_base_wis = standardization(Base_de_données_wisconsin)

base_test_wisconsin = Base_de_données_wisconsin[400:]
base_donnée_wisconsin = Base_de_données_wisconsin [:400]


# base_donnéesur1_wis = sur1(base_donnée_wisconsin,8)
# base_testsur1_wis = sur1 ( base_test_wisconsin,8)

stand_donnée_wis = stand_base_wis [:400]
stand_test_wis =  stand_base_wis [400:]

def proportion_wis (dataset,nbr_voisins):
    resultat_normal = []
    resultat_distance = []
    base = standardization(dataset)
    a = base[400:]
    b = base [:400]
    for k in range (1,401):
        y = b[:k]
        c , d = resultat_test(a,y,nbr_voisins)
        resultat_normal.append(d)
        e,f = resultat_test_distance(a,y,nbr_voisins)
        resultat_distance.append(f)
    x = [ k for k in range (1,401)]
    plt.plot(x,resultat_normal,'r',label='normal')
    plt.plot(x,resultat_distance,'b',label='distance')
    plt.show()
    return (resultat_normal,resultat_distance)


## Tableau recapitulatif

def tableau_proches (a,dataset):
    with open ('tableau_proches.csv','w',newline='') as f :
        thewriter = csv.writer(f,delimiter = ';')

        thewriter.writerow(a + [prediction(a,dataset)])
        thewriter.writerow(" ")
        proches = [plus_proches(a,dataset)[k][0] for k in range(len(plus_proches(a,dataset)))]
        for k in proches :
            thewriter.writerow([k] + dataset[k])

def write_column (a):
    with open ('tableau_rendu.csv' , 'w' , newline = '' ) as f :
        writer = csv.writer(f,delimiter=';')

        t = []
        for k in range (len(a[0])):
            t.append([a[l][k] for l in range (len(a)) ] )
        for k in range (len(t)):
            writer.writerow(t[k])

# On obtient bien les proches en colonne rangés par catégorie diagnostic
# on considere le rang dans le tableau traité comme l'Id des patients

def tableau_rendu (stand_a,a,dataset,stand_dataset,nb_voisins) :

    with open ('tableau_rendu.csv' , 'w' , newline = '' ) as f :
        writer = csv.writer(f,delimiter=';')
        writer.writerow('')

    # on veut les valeurs des mesures pour que le medecin puisse les interpréter
    # et pas les valeurs standardizées pour les calculs
        proches = [plus_proches(stand_a,stand_dataset,nb_voisins)[k][0] for k in range(len(plus_proches(stand_a,stand_dataset,nb_voisins)))]
        T = [[' ',' ',k,' ',' '] + dataset[k] for k in proches]
        B = [T[k][:-1] for k in range (len(T))  if T[k][-1] == 'B']
        M = [T[k][:-1] for k in range (len(T))  if T[k][-1] == 'M']
        vide = [ '' for k in range (len(a)+5)]

        L = B + [vide] + [ [' ',prediction_distance(stand_a,stand_dataset,nb_voisins),'Patient',' ',' '] + a] + [vide] + M
        write_column(L)
        return L

## Base de données 2

# Implémentation d'une deuxieme base de données de taille 830
def lecture_fichier2():
    with open ('mammographic_mass.csv','r') as base_donnee :
        csv_reader = csv.reader(base_donnee)
        dataset = []
        for line in (base_donnee):
            dataset.append(line.split(","))
        for k in range (1,len(dataset)):
            for l in range (len(dataset[k])):
                dataset[
                k][l] = float(dataset[k][l])
        return dataset[1:]

dataset_mammographic_mass = lecture_fichier2()

def mise_en_forme2(dataset):
    data = [ dataset[k][:-1] + [ 'B' if dataset[k][-1] == 0 else 'M' ] for k in range (len(dataset)) ]
    return data

base_de_données_mass_originale = mise_en_forme2(dataset_mammographic_mass)
base_de_données_mass = shuffle(base_de_données_mass_originale)
stand_base_mass = standardization(base_de_données_mass)

base_test_mass = base_de_données_mass[600:]
base_donnée_mass = base_de_données_mass [:600]


# donnée_mass_sur1 = sur1(base_donnée_mass,8)
# test_mass_sur1 = sur1(base_test_mass,8)


stand_donnée_mass = stand_base_mass[:400]
stand_test_mass = stand_base_mass[400:]



## tests

def test_wis (a,nbr_voisins):
    stand_donnée_wis = stand_base_wis[:a]
    stand_test_wis = stand_base_wis[a:]
    x = resultat_test(stand_test_wis,stand_donnée_wis,nbr_voisins)
    y = resultat_test_distance(stand_test_wis,stand_donnée_wis,nbr_voisins)
    return x,y

from math import *

def proportions_mass(dataset,nbr_voisins):
    resultat_normal = []
    resultat_distance = []
    a = base_de_données_mass[500:]
    b = base_de_données_mass [:500]
    for k in range (1,501):
        y = b[:k]
        c , d = resultat_test(a,y,nbr_voisins)
        resultat_normal.append(d)
        e,f = resultat_test_distance(a,y,nbr_voisins)
        resultat_distance.append(f)
    x = [ k for k in range (1,501)]
    plt.plot(x,resultat_normal,'r',label='normal')
    plt.plot(x,resultat_distance,'b',label='distance')
    plt.show()
    return (resultat_normal,resultat_distance)

def test_mass (a,nbr_voisins):
    stand_donnée_mass = stand_base_mass[:a]
    stand_test_mass = stand_base_mass[a:]
    x = resultat_test(stand_test_mass,stand_donnée_mass,nbr_voisins)
    y = resultat_test_distance(stand_test_mass,stand_donnée_mass,nbr_voisins)
    return x,y

def approximation_percentage_wis (dataset) :
    x = [ k for k in range (1,251)]
    y = [75 for k in range (1,38)]+[ 800/(k) for k in range(11,251-27) ]
    plt.plot(x,y,'b')
    alpha,betha = proportions_wis(dataset)
    plt.plot(x,alpha,'r')
    plt.plot(x,betha,'b')
    plt.show()



## reduc dimension scikit
from sklearn.decomposition import PCA

def localisation_erreurs_pca (test,dataset,predict,nbr_voisins):
    a = datatest(test)
    b = reponse_test(test)
    c = [predict(test[k][:-1],dataset,nbr_voisins) for k in range (len(test))]

    d = datatest(dataset)
    e = reponse_test(dataset)


    pca = PCA(n_components=2)
    principalComponents = pca.fit_transform(a)
    Y = principalComponents.tolist()

    principalComponents2 = pca.fit_transform(d)
    Y2 = principalComponents2.tolist()

    fig1 = plt.figure(1)
    ax1 = fig1.gca()



    for i in range (len(Y2)) :
        if e[i] == 'M' :
            ax1.scatter(Y2[i][0],Y2[i][1],c='r')
        else : ax1.scatter(Y2[i][0],Y2[i][1],c='b')

    for k in range (len(Y)):
        if b[k] == 'M' and c[k] == 'B' :
            ax1.scatter(Y[k][0],Y[k][1],c='k')
        elif  b[k] == 'B' and c[k] == 'M' :
            ax1.scatter(Y[k][0],Y[k][1],c='g')

    plt.show()

def visualisation_situation (a,dataset,nb_voisins):

    d = datatest(dataset)+[a]
    e = reponse_test(dataset)
    plus_proche = plus_proches(a,dataset,nb_voisins)
    proches = [plus_proche[k][0] for k in range(len(plus_proche))]

    pca = PCA(n_components=2)
    principalComponents2 = pca.fit_transform(d)
    Y = principalComponents2.tolist()
    data_pca = Y[:-1]
    a_pca = Y[-1]

    fig1 = plt.figure(1)
    ax1 = fig1.gca()

    for i in proches :
        if e[i] == 'M' :
            ax1.scatter(data_pca[i][0],data_pca[i][1],c='r')
        else : ax1.scatter(data_pca[i][0],data_pca[i][1],c='b')
    ax1.scatter(a_pca[0],a_pca[1],c='k')

    u,v = a_pca
    #cercle(u,v,0.35)
    #cercle(u,v,1.4)
    plt.axis('equal')
    plt.show()

def cercle (a,b,r):
    n = 1000
    x = [a+r*cos(2*pi*k/n) for k in range (0,n)]
    y = [b+r*sin(2*pi*k/n) for k in range (0,n)]
    plt.plot(x,y,'k')

## PCA fait maison


def mean (dataset):
    l = len(dataset)
    p = len(dataset[0])
    mean = [0 for i in range (p)]
    for k in range (l):
        for i in range (p):
            mean[i] += dataset[k][i]
    for i in range (p):
        mean[i] = mean[i]/l
    return mean

def variance (dataset,i,mean):
    l = len(dataset)
    res = 0
    mean_i = mean[i]
    for k in range(l):
        res += (dataset[k][i]-mean_i)*(dataset[k][i]-mean_i)
    return res/l

def covariance (dataset,i,j,liste_mean):
    l = len(dataset)
    res = 0
    mean_i = liste_mean[i]
    mean_j = liste_mean[j]
    for k in range(l):
        res += (dataset[k][i]-mean_i)*(dataset[k][j]-mean_j)
    return res/l

def matrice_covariance (dataset):
    liste_mean = mean(dataset)
    p = len(liste_mean)
    matrice_covariance = []
    for k in range (p):
        m = []
        for l in range (k+1):
            m.append(covariance(dataset,k,l,liste_mean))
        comble = [0 for i in range (p-k-1)]
        matrice_covariance.append(m+comble)
    for k in range (p-1):
        for l in range (k+1,p):
            matrice_covariance[k][l] = matrice_covariance[l][k]
    return matrice_covariance

def new_matrice(M,l):
    new_M = M[::]
    for k in range(len(new_M)):
        new_M[k]=new_M[k][:l]+new_M[k][l+1:]
    return new_M[1:]

def determinant_matrice(M):
    l = len(M)
    if l == 1 : return M[0][0]
    else :
        sum = 0
        signe = 1
        # developpement sur la première ligne
        for k in range(l):
            new_M = new_matrice(M,k)
            sum += M[0][k] * signe * determinant_matrice(new_M)
            signe = -signe
        return sum

def eigen(M):
    matrice = np.array(M)
    eigen = np.linalg.eig(matrice)
    return eigen

from test_val_propres import *
def eigen_maison(M):
    return vects_propres(M)

def tri_eigen(Eigen):
    eigen_values , eigen_vectors =  Eigen
    eigen_values = eigen_values.tolist()
    eigen_vectors = eigen_vectors.tolist()
    sorted_values = []
    sorted_vectors = []
    for k in range(len(eigen_values)):
        m = eigen_values.index(max(eigen_values))
        sorted_values.append(eigen_values.pop(m))
        sorted_vectors.append(eigen_vectors.pop(m))
    return sorted_values,sorted_vectors

def dot_product (a,b):
    res = 0
    l = len(a)
    for k in range (len(a)):
        res += a[k]*b[k]
    return res

def mult_vect (scalaire,vect):
    return list(map(lambda x : scalaire*x , vect))

def add_vect (vect1,vect2):
    return list(map(lambda x,y : x+y , vect1, vect2))


def projection (eigen_vectors,dimension,dataset):
    data = dataset[::]
    vecteurs = eigen_vectors[:dimension]
    for k in range (len(dataset)):
        new_point = [0 for k in range (dimension)]
        for i in range (dimension):
            eigen_i = vecteurs[i]
            dot_prod = dot_product(dataset[k],eigen_i)
            new_point = add_vect(new_point,mult_vect(dot_prod,eigen_i))
        data[k] = new_point
    return data

def PCA_maison(dataset,dimension):
    M_cov = matrice_covariance(dataset)
    vectors = tri_eigen(eigen(M_cov))[1]
    return projection(vectors,dimension,dataset)

def trace_PCA_2D(dataset):
    data = PCA_maison(dataset,2)
    x = [data[k][0] for k in range (len(dataset))]
    y = [data[k][1] for k in range (len(dataset))]
    plt.scatter(x,y)
    plt.show()

def trace_couleur_pca_2D(dataset):
    data = datatest(dataset)
    pca = PCA_maison(data,2)
    rep = reponse_test(dataset)
    for k in range (len(dataset)):
        if rep[k] == 'B':
            plt.scatter(pca[k][0],pca[k][1],c='b')
        else : plt.scatter(pca[k][0],pca[k][1],c='r')
    plt.show()

def trace_couleur_pca_1D(dataset):
    data = datatest(dataset)
    pca = PCA_maison(data,1)
    rep = reponse_test(dataset)
    for k in range (len(dataset)):
        if rep[k] == 'B':
            plt.scatter(k,pca[k][0],c='b')
        else : plt.scatter(k,pca[k][0],c='r')
    plt.show()

def trace_couleur_pca_3D(dataset):
    data = datatest(dataset)
    pca = PCA_maison(data,3)
    rep = reponse_test(dataset)
    ax = plt.axes(projection='3d')
    for k in range (len(dataset)):
        if rep[k] == 'B':
            ax.scatter3D(pca[k][0],pca[k][1],pca[k][2],c='b')
        else : ax.scatter3D(pca[k][0],pca[k][1],pca[k][2],c='r')
    plt.show()

from mpl_toolkits import mplot3d
import numpy as np

def localisation_erreurs_pca_maison (test,dataset,predict,nbr_voisins):
    a = datatest(test)
    b = reponse_test(test)
    c = [predict(test[k][:-1],dataset,nbr_voisins) for k in range (len(test))]

    d = datatest(dataset)
    e = reponse_test(dataset)


    pca_test = PCA_maison(a,2)
    Y = pca_test

    pca_dataset = PCA_maison(d,2)
    Y2 = pca_dataset

    fig1 = plt.figure(1)
    ax1 = fig1.gca()



    for i in range (len(Y2)) :
        if e[i] == 'M' :
            ax1.scatter(Y2[i][0],Y2[i][1],c='r')
        else : ax1.scatter(Y2[i][0],Y2[i][1],c='b')

    for k in range (len(Y)):
        if b[k] == 'M' and c[k] == 'B' :
            ax1.scatter(Y[k][0],Y[k][1],c='k')
        elif  b[k] == 'B' and c[k] == 'M' :
            ax1.scatter(Y[k][0],Y[k][1],c='g')

    plt.show()

def visualisation_situation_maison (a,dataset,nb_voisins):


    d = datatest(dataset)+[a]
    e = reponse_test(dataset)
    plus_proche = plus_proches(a,dataset,nb_voisins)
    proches = [plus_proche[k][0] for k in range(len(plus_proche))]

    pca = PCA_maison(d,2)
    Y = pca
    data_pca = Y[:-1]
    a_pca = Y[-1]

    fig1 = plt.figure(1)
    ax1 = fig1.gca()

    for i in proches :
        if e[i] == 'M' :
            ax1.scatter(data_pca[i][0],data_pca[i][1],c='r')
        else : ax1.scatter(data_pca[i][0],data_pca[i][1],c='b')
    ax1.scatter(a_pca[0],a_pca[1],c='k')

    u,v = a_pca
    #cercle(u,v,0.35)
    #cercle(u,v,1.4)
    plt.axis('equal')
    plt.show()

def heatmap_correlation(dataset):
    cov_data = np.array(matrice_covariance(dataset))
    img = plt.matshow(cov_data, cmap=plt.cm.rainbow)
    plt.colorbar(img, ticks = [-1, 0, 1], fraction=0.045)
    for x in range(cov_data.shape[0]):
        for y in range(cov_data.shape[1]):
            plt.text(x, y, "%0.2f" % cov_data[x,y], size=10, color='black', ha="center", va="center")

    plt.show()

## Interface

from tkinter import *
import os
import time


def interface_mass_knn():


    def onclick_mass():
        patient = [bi_rads.get(),age.get(),forme.get(),marge.get(),densite.get()]
        data = base_de_données_mass[::]
        data.append(patient)
        stand_data = standardization(data)
        stand_patient = stand_data[-1][:-1]
        stand_base_donnée_mass = stand_data[:-1]

        nb_voisins = nb_mass

        visualisation_situation(stand_patient,stand_base_donnée_mass,nb_voisins)
        tableau_rendu(stand_patient,patient,base_de_données_mass,stand_base_mass,nb_voisins)
        coloriage_tableau_rendu()
        os.system('tableau_colorie.xlsx')

        predict = prediction(stand_patient,stand_base_donnée_mass,nb_voisins)
        percentage = percentage_prediction_distance(stand_patient,stand_base_donnée_mass,nb_voisins)
        [(predict,perc_predict),(autre,perc_autre)] = percentage


        window2 = Tk()
        window2.title("Breast Cancer Predictor")
        window2.geometry("550x400")
        window2.minsize(550,400)
        window2.iconbitmap("logo_predictor.ico")
        window2.config(background='#F6C0DD')

        frame_globale2 = Frame(window2,bg='#F6C0DD')
        frame_percentage = Frame(frame_globale2,bg = '#C56D9C')

        if predict == 'B' :
            Label(frame_globale2, text="Le diagnostic du patient est négatif", font=("arial", 20), fg="white", bg="#F6C0DD", height=2).pack(expand = YES)
        else :
            Label(frame_globale2, text="Le diagnostic du patient est positif", font=("arial", 20), fg="white", bg="#F6C0DD", height=2).pack(expand = YES)

        texte_predict = Label(frame_percentage, text = predict,pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 0)
        pourcentage_predict = Label(frame_percentage, text = (str(100*perc_predict)+'%')[:7],pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 0)
        texte_autre = Label(frame_percentage, text = autre,pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 1)
        pourcentage_autre = Label(frame_percentage, text = (str(100*perc_autre)+'%')[:7],pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 1)

        frame_percentage.pack(expand=YES)

        frame_globale2.pack(expand = YES)

    # créer fenetre
    window = Tk()

    # pesonnalise fenetre
    window.title("Breast Cancer Predictor KNN Mass")
    window.geometry("1080x720")
    window.minsize(480,360)
    window.iconbitmap("logo_predictor.ico")
    window.config(background='#F6C0DD')

    # créer frame

    frame_globale = Frame(window,bg='#F6C0DD')
    frame_données = Frame(frame_globale,bg = '#C56D9C')


    # ajout texte global
    texte_globale = Label(frame_globale, text = "Veuillez rentrer les données",pady = 15, font = ('arial', 20) , bg = '#F6C0DD' ,fg = '#FFFFFF')
    texte_globale.pack()

    # ajout texte frame données
    texte_donnée1 = Label(frame_données, text = "BI-Rades",pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 0)
    bi_rads = DoubleVar()
    Entry(frame_données,textvariable= bi_rads,width=10, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=0,column=1)

    texte_donnée2 = Label(frame_données, text = "Age",pady = 10,padx = 50, font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 0)
    age = DoubleVar()
    Entry(frame_données,textvariable= age,width=10, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=1,column=1)

    texte_donnée3 = Label(frame_données, text = "Forme",pady = 10,padx = 50, font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 2 , column = 0)
    forme = DoubleVar()
    Entry(frame_données,textvariable= forme,width=10, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=2,column=1)

    texte_donnée4 = Label(frame_données, text = "Marge",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 3 , column = 0)
    marge = DoubleVar()
    Entry(frame_données,textvariable= marge,width=10, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=3,column=1)

    texte_donnée5 = Label(frame_données, text = "Densité",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 4 , column = 0)
    densite = DoubleVar()
    Entry(frame_données,textvariable= densite,width=10, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=4,column=1)

    #ajouter bouton

    predict_button = Button(frame_globale,text = 'Prédiction', font = ('arial', 20) , bg = '#C56D9C', fg = '#FFFFFF',command = onclick_mass)
    predict_button.pack(pady = 20, side = BOTTOM , fill = X)

    frame_données.pack(expand=YES,fill = X)

    frame_globale.pack(expand = YES)

    window.mainloop()



def interface_wisconsin_knn():


    def onclick_wisconsin():
        patient=[radius_mean.get(),texture_mean.get(),perimeter_mean.get(),area_mean.get(),smoothness_mean.get(),compactness_mean.get(),concavity_mean.get(),concave_points_mean.get(),symmetry_mean.get(),fractal_dimension_mean.get(),radius_se.get(),texture_se.get(),perimeter_se.get(),area_se.get(),smoothness_se.get(),compactness_se.get(),concavity_se.get(),concave_points_se.get(),symmetry_se.get(),fractal_dimension_se.get(),radius_worst.get(),texture_worst.get(),perimeter_worst.get(),area_worst.get(),smoothness_worst.get(),compactness_worst.get(),concavity_worst.get(),concave_points_worst.get(),symmetry_worst.get(),fractal_dimension_worst.get()]

        data = Base_de_données_wisconsin[::]
        data.append(patient)
        stand_data = standardization(data)
        stand_patient = stand_data[-1][:-1]

        stand_base_donnée_wis = stand_data[:-1]

        nb_voisins = nb_wis

        visualisation_situation(stand_patient,stand_base_donnée_wis,nb_voisins)
        tableau_rendu(stand_patient,patient,Base_de_données_wisconsin,stand_base_wis,nb_voisins)
        coloriage_tableau_rendu()
        os.system('tableau_colorie.xlsx')

        prediction=prediction_distance(stand_patient,stand_base_donnée_wis,nb_voisins)
        percentage = percentage_prediction_distance(stand_patient,stand_base_donnée_wis,nb_voisins)
        [(predict,perc_predict),(autre,perc_autre)] = percentage

        window2 = Tk()
        window2.title("Breast Cancer Predictor")
        window2.geometry("550x400")
        window2.minsize(550,400)
        window2.iconbitmap("logo_predictor.ico")
        window2.config(background='#F6C0DD')

        frame_globale2 = Frame(window2,bg='#F6C0DD')
        frame_percentage = Frame(frame_globale2,bg = '#C56D9C')


        if prediction == 'B' :
            Label(frame_globale2, text="Le diagnostic du patient est négatif", font=("arial", 20), fg="white", bg="#F6C0DD", height=2).pack(expand = YES)
        else :
            Label(frame_globale2, text="Le diagnostic du patient est positif", font=("arial", 20), fg="white", bg="#F6C0DD", height=2).pack(expand = YES)

        texte_predict = Label(frame_percentage, text = predict,pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 0)
        pourcentage_predict = Label(frame_percentage, text = (str(100*perc_predict)+'%')[:7],pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 0)
        texte_autre = Label(frame_percentage, text = autre,pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 1)
        pourcentage_autre = Label(frame_percentage, text = (str(100*perc_autre)+'%')[:7],pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 1)

        frame_percentage.pack(expand=YES)

        frame_globale2.pack(expand = YES)

    # créer fenetre
    window = Tk()

    # pesonnalise fenetre
    window.title("Breast Cancer Predictor KNN Wisconsin")
    width= window.winfo_screenwidth()
    height= window.winfo_screenheight()
    window.geometry('%dx%d' % (width, height))
    window.minsize(1000,800)
    window.iconbitmap("logo_predictor.ico")
    window.config(background='#F6C0DD')

    # créer frame

    frame_globale = Frame(window,bg='#F6C0DD')
    frame_données = Frame(frame_globale,bg = '#C56D9C')

    # ajout texte global
    texte_globale = Label(frame_globale, text = "                                                                 Veuillez rentrer les données                                                                   ",pady = 15, font = ('arial', 20) , bg = '#F6C0DD' ,fg = '#FFFFFF')
    texte_globale.pack()

    # ajout texte frame données
    texte_donnée1 = Label(frame_données, text = "Rayon moyen",pady = 10,padx = 50, font = ('arial', 14) , bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 0)
    radius_mean = DoubleVar()
    Entry(frame_données,textvariable= radius_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=0,column=1)

    texte_donnée2 = Label(frame_données, text = "Texture moyenne",pady = 10,padx = 50, font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 0)
    texture_mean = DoubleVar()
    Entry(frame_données,textvariable= texture_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=1,column=1)

    texte_donnée3 = Label(frame_données, text = "Périmètre moyen",pady = 10,padx = 50, font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 2 , column = 0)
    perimeter_mean = DoubleVar()
    Entry(frame_données,textvariable= perimeter_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=2,column=1)

    texte_donnée4 = Label(frame_données, text = "Aire moyenne",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 3 , column = 0)
    area_mean = DoubleVar()
    Entry(frame_données,textvariable= area_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=3,column=1)

    texte_donnée5 = Label(frame_données, text = "Lisseté moyenne",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 4 , column = 0)
    smoothness_mean = DoubleVar()
    Entry(frame_données,textvariable= smoothness_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=4,column=1)

    texte_donnée6 = Label(frame_données, text = "Compacité moyenne",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 5 , column = 0)
    compactness_mean = DoubleVar()
    Entry(frame_données,textvariable= compactness_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=5,column=1)

    texte_donnée7 = Label(frame_données, text = "Concavité moyenne",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 6 , column = 0)
    concavity_mean = DoubleVar()
    Entry(frame_données,textvariable= concavity_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=6,column=1)

    texte_donnée8 = Label(frame_données, text = "Nombre moyen de creux",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 7 , column = 0)
    concave_points_mean = DoubleVar()
    Entry(frame_données,textvariable= concave_points_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=7,column=1)

    texte_donnée9 = Label(frame_données, text = "Symétrie moyenne",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 8 , column = 0)
    symmetry_mean = DoubleVar()
    Entry(frame_données,textvariable= symmetry_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=8,column=1)

    texte_donnée10 = Label(frame_données, text = "Dimension fractale moyenne",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 9 , column = 0)
    fractal_dimension_mean = DoubleVar()
    Entry(frame_données,textvariable= fractal_dimension_mean,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=9,column=1)

    texte_donnée11 = Label(frame_données, text = "Ecart-type Rayon",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 2)
    radius_se = DoubleVar()
    Entry(frame_données,textvariable= radius_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=0,column=3)

    texte_donnée12 = Label(frame_données, text = "Ecart-type Texture",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 2)
    texture_se = DoubleVar()
    Entry(frame_données,textvariable= texture_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=1,column=3)

    texte_donnée13 = Label(frame_données, text = "Ecart-type Perimetre",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 2 , column = 2)
    perimeter_se = DoubleVar()
    Entry(frame_données,textvariable= perimeter_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=2,column=3)

    texte_donnée14 = Label(frame_données, text = "Ecart-type Aire",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 3 , column = 2)
    area_se = DoubleVar()
    Entry(frame_données,textvariable= area_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=3,column=3)

    texte_donnée15 = Label(frame_données, text = "Ecart-type Lisseté",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 4 , column = 2)
    smoothness_se = DoubleVar()
    Entry(frame_données,textvariable= smoothness_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=4,column=3)

    texte_donnée16 = Label(frame_données, text = "Ecart-type Compacité",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 5 , column = 2)
    compactness_se = DoubleVar()
    Entry(frame_données,textvariable= compactness_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=5,column=3)

    texte_donnée17 = Label(frame_données, text = "Ecart-type Concavité",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 6 , column = 2)
    concavity_se = DoubleVar()
    Entry(frame_données,textvariable= concavity_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=6,column=3)

    texte_donnée18 = Label(frame_données, text = "Ecart-type Nombre points concaves",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 7 , column = 2)
    concave_points_se = DoubleVar()
    Entry(frame_données,textvariable= concave_points_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=7,column=3)

    texte_donnée19 = Label(frame_données, text = "Ecart-type Symétrie",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 8 , column = 2)
    symmetry_se = DoubleVar()
    Entry(frame_données,textvariable= symmetry_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=8,column=3)

    texte_donnée20 = Label(frame_données, text = "Ecart-type Dimension fractale",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 9 , column = 2)
    fractal_dimension_se = DoubleVar()
    Entry(frame_données,textvariable= fractal_dimension_se,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=9,column=3)

    texte_donnée21 = Label(frame_données, text = "Pire Rayon",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 0 , column = 4)
    radius_worst = DoubleVar()
    Entry(frame_données,textvariable= radius_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=0,column=5)

    texte_donnée22 = Label(frame_données, text = "Pire Texture",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 1 , column = 4)
    texture_worst = DoubleVar()
    Entry(frame_données,textvariable= texture_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=1,column=5)

    texte_donnée23 = Label(frame_données, text = "Pire Périmètre",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 2 , column = 4)
    perimeter_worst = DoubleVar()
    Entry(frame_données,textvariable= perimeter_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=2,column=5)

    texte_donnée24 = Label(frame_données, text = "Pire Aire",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 3 , column = 4)
    area_worst = DoubleVar()
    Entry(frame_données,textvariable= area_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=3,column=5)

    texte_donnée25 = Label(frame_données, text = "Pire Lisseté",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 4 , column = 4)
    smoothness_worst = DoubleVar()
    Entry(frame_données,textvariable= smoothness_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=4,column=5)

    texte_donnée26 = Label(frame_données, text = "Pire Compacité",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 5 , column = 4)
    compactness_worst = DoubleVar()
    Entry(frame_données,textvariable= compactness_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=5,column=5)

    texte_donnée27 = Label(frame_données, text = "Pire Concavité",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 6 , column = 4)
    concavity_worst = DoubleVar()
    Entry(frame_données,textvariable= concavity_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=6,column=5)

    texte_donnée28 = Label(frame_données, text = "Pire Nombre points concaves",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 7 , column = 4)
    concave_points_worst = DoubleVar()
    Entry(frame_données,textvariable= concave_points_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=7,column=5)

    texte_donnée29 = Label(frame_données, text = "Pire Symétrie",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 8 , column = 4)
    symmetry_worst = DoubleVar()
    Entry(frame_données,textvariable= symmetry_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=8,column=5)

    texte_donnée30 = Label(frame_données, text = "Pire Dimension fractale",pady = 10,padx = 50,font=('arial', 14),bg = '#C56D9C', fg = '#FFFFFF').grid(row = 9 , column = 4)
    fractal_dimension_worst = DoubleVar()
    Entry(frame_données,textvariable= fractal_dimension_worst,width=7, font = ('arial',15) , bg = '#F6C0DD',fg = '#FFFFFF').grid(row=9,column=5)




    #ajouter bouton

    predict_button = Button(frame_globale,text = 'Prédiction', font = ('arial', 20) , bg = '#C56D9C', fg = '#FFFFFF',command = onclick_wisconsin)
    predict_button.pack(pady = 20, side = BOTTOM , fill = X)


    # affichage

    frame_données.pack(expand=YES,fill = X)

    frame_globale.pack(expand = YES)

    window.mainloop()


## mise en forme tableau rendu ( coloriage )

from xlsxwriter.workbook import Workbook

def csv_to_xlsx (nom) :
    csvfile = nom
    workbook = Workbook(csvfile[:-4] + '.xlsx')
    worksheet = workbook.add_worksheet()
    f = open(csvfile, 'r', encoding='utf8')
    liste = f.readlines()
    liste_final = []
    for ligne in liste :
        liste_final.append(ligne.strip().split(";"))



    for k in range (2,len(liste_final)):
        for j in range (len(liste_final[2])):
            a = liste_final[k][j]
            try : texte = float(a)
            except : texte = a
            worksheet.write(k, j, texte )

    workbook.close()


import openpyxl

def lecture_couleur_case (nom,i,j):
    workbook = openpyxl.load_workbook(nom)
    sheet = workbook.active
    fill = str(sheet.cell(i,j).fill)
    new_f = fill.strip().split(",")
    f = new_f [1]
    couleur = f.split("\n")[-1]
    code_couleur = couleur[-7:-1]
    workbook.close()
    return(code_couleur)

def palette_couleur():
    nom = "tab_couleur.xlsx"
    vert = lecture_couleur_case(nom,1,1)
    rouge = lecture_couleur_case(nom,1,11)
    palette_couleur = [("vert",vert),("rouge",rouge)]
    degrade_bleu = []
    degrade_rouge = []
    for k in range (1,8):
        degrade_bleu.append(lecture_couleur_case(nom,3,k))
        degrade_rouge.append(lecture_couleur_case(nom,3,10+k))
    palette_couleur.append(("degrade_bleu",degrade_bleu))
    palette_couleur.append(("degrade_rouge",degrade_rouge))
    return palette_couleur

def localisation_patient(nom):
    workbook = openpyxl.load_workbook(nom)
    sheet = workbook.active
    max_column = sheet.max_column
    for k in range (1,max_column + 1):
        if sheet.cell(4,k).value == "Patient":
            return k
    workbook.close()
    return ("le patient n'est pas présent'")

def coloriage_bleu (nom):
    indice_patient = localisation_patient(nom) -1
    workbook = openpyxl.load_workbook(nom)
    sheet = workbook.active
    max_row = sheet.max_row
    max_column = sheet.max_column
    palette = palette_couleur()


    for i in range(7,max_row+1):
        row = []
        if indice_patient-1 > 0 :
            for j in range (1,indice_patient):
                row.append((sheet.cell(i,j).value))

            donnée_patient = sheet.cell(i,indice_patient + 1).value
            ecarts = ecart_relatif(donnée_patient,row)
            couleur = couleurs(ecarts)

            for j in range (1,indice_patient):
                indice_couleur = couleur[j-1]

                color = palette[2][1][indice_couleur]
                Fill = PatternFill(fgColor= color, bgColor="FFFFFF", fill_type='solid')
                cell = sheet.cell(i,j)
                cell.fill = Fill

    for i in range(7,max_row+1):
        row = []
        if max_column - indice_patient -1 > 0 :
            for j in range (indice_patient + 3 , max_column + 1):
                row.append((sheet.cell(i,j).value))

            donnée_patient = sheet.cell(i,indice_patient + 1).value
            ecarts = ecart_relatif(donnée_patient,row)
            couleur = couleurs(ecarts)

            for j in range (indice_patient + 3 , max_column + 1):
                indice_couleur = couleur[j -(indice_patient + 3) ]

                color = palette[3][1][indice_couleur]
                Fill = PatternFill(fgColor= color, bgColor="FFFFFF", fill_type='solid')
                cell = sheet.cell(i,j)

                cell.fill = Fill

    for i in range (3,5):
        for j in range (1,max_column +1):
            if j < indice_patient :
                color = palette[0][1]
                Fill = PatternFill(fgColor= color, bgColor="FFFFFF", fill_type='solid')
                cell = sheet.cell(i,j)
                cell.fill = Fill

            elif j > indice_patient + 2 :
                color = palette[1][1]
                Fill = PatternFill(fgColor= color, bgColor="FFFFFF", fill_type='solid')
                cell = sheet.cell(i,j)
                cell.fill = Fill


    workbook.save('tableau_colorie.xlsx')


def coloriage_tableau_rendu():
    csv_to_xlsx("tableau_rendu.csv")
    coloriage_bleu("tableau_rendu.xlsx")





from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font, Color



liste_dégradé = [0.08+k*0.08 for k in range (7)]

def ecart_relatif(patient,liste):
    ecart = []
    for k in range (len(liste)):
        if patient != 0 :
            ecar = abs(liste[k]-patient)/patient
        else : ecar = liste[k]/100
        ecart.append (ecar)
    return ecart


def couleurs(liste):
    couleurs = []
    for k in range (len(liste)):
        i = 0
        while i<6 and liste[k] > liste_dégradé[i] :
            i += 1
        couleurs.append(i)
    return couleurs

## courbes densite/repartion valeurs dataset

from scipy.integrate import simps
from scipy.ndimage import gaussian_filter1d

import matplotlib

def courbes_param(dataset,n):
    gauss_B = []
    gauss_M = []
    area_B = []
    area_M = []
    liste_h = []

    for k in range(len(dataset[0])-1):
        param = [donnée[k] for donnée in dataset]
        maxi = max(param)
        mini = min(param)
        repartition_B = [0 for k in range(n)]
        repartition_M = [0 for k in range(n)]
        b = 0
        m = 0
        h = (maxi - mini)/n
        liste_h.append(h)

        for i in range(len(param)) :
            valeur = param[i]
            diagnostic = dataset[i][-1]
            indice = int((valeur - mini)// h)-1


            if diagnostic =='B' :
                repartition_B[indice] = repartition_B[indice]+1
                b += 1
            else :
                repartition_M[indice] = repartition_M[indice] + 1
                m += 1

        for i in range(len(repartition_B)) :
            repartition_B[i] = repartition_B[i]/b
        for i in range(len(repartition_M)) :
            repartition_M[i] = repartition_M[i]/m


        x = [mini + k*h for k in range (n) ]
        y_gauss_B = gaussian_filter1d(repartition_B, sigma=5)
        y_gauss_M = gaussian_filter1d(repartition_M, sigma=5)
        gauss_B.append(y_gauss_B)
        gauss_M.append(y_gauss_M)

        area_B_k = simps(y_gauss_B, dx= h )
        area_M_k = simps(y_gauss_M, dx= h )
        area_B.append(area_B_k)
        area_M.append(area_M_k)

        plt.figure(k)
        plt.plot(x,y_gauss_B,c='b')
        plt.plot(x,y_gauss_M,c='r')
        plt.fill_between(x, y_gauss_B, color='#539ecd')
        plt.fill_between(x, y_gauss_M, color='#FF7F50')

        #plt.savefig('image plot/courbes densite param wis/courbe_param_{}'.format(k))
        plt.close()
    return [gauss_B,gauss_M,area_B,area_M,liste_h]


def scatter_pram(dataset):
    for k in range (len(dataset[0])-1):
        for j in range(k+1):
            param_k = [donnée[k] for donnée in dataset]
            param_j = [donnée[j] for donnée in dataset]
            plt.figure(k)

            for i in range (len(dataset)):
                if dataset[i][-1] == 'B' : color = 'b'
                else : color = 'r'
                x = param_k[i]
                y = param_j[i]

                plt.scatter(x,y,c=color)

            plt.savefig('image plot/scatter_param_mass/scatter_param_mass_{},{}'.format(k,j))
            plt.close()

# Variance inflation factor to select features

## selection paramètres

# On cherche tout d'abord à éliminer les données doubles, c'est à dire les paramètres ayant une grande corrélation entre eux car ils nous apportent en fait les mêmes données et on augmente le temps de calcul.

# Parmi les paramètres restant, ils n'ont pas tous la même importance, mais comment définir l'importance ?

# Ici ce qui nous intéresse c'est la faculté à déterminer la classe en fonction de la valeur d'un paramètre donnée, ainsi, en s'apuyant sur les courbes de densité/répartion des valeurs, on peut essayer de calculer à quel point cette dissociation est possible.
# deux courbes qui sont identiques => impossible de discerner,
# deux courbes séparées par un creux => parfait

# Ainsi, on imagine une droite verticale, et on calcul l'aire sous la courbe bleu à gauche de la droite, celle de la courbe rouge à droite de la droite et on fait la proportion aire_B_gauche/aire_B_totale de meme pour rouge, on les multiplie, et on cherche a avoir le meilleur score possible car avoir un score élevé c'est arriver à séparer efficacement les malades des personnes saines, on ne cherche pas à déterminer la valeur seuil, juste cette efficacité

# Un paramètre selon lequel M et B sont discernable => paramètre important pour la classification, on peut donc alors coefficienter les distances paramètre par paramètre pour une classification plus complexe et en théorie plus précise.


def elimine_param(liste_param,dataset):
    res = []
    for k in range (len(dataset)):
        donnée = []
        for j in range(len(dataset[0])):
            if j not in liste_param :
                donnée.append(dataset[k][j])
        res.append(donnée)
    return res

# en éliminant les paramètres ayant une corrélation tq |cor(x,y)| > 0.80
liste_param_elim_80 = [2,3,6,7,12,13,16,17,20,21,22,23,24,25,26,27]
liste_param_elim_90 = [2,3,6,12,13,20,21,22,23,27]

stand_base_wis_clean80 = elimine_param(liste_param_elim_80,stand_base_wis)
stand_base_wis_clean90 = elimine_param(liste_param_elim_90,stand_base_wis)

# performances des données avec eval perf, n=200 sur la base wis :
#                           mean perf normale        mean perf distance   nombre de paramètres

# performances avec tout :   (3.9204545454545463, 3.3216374269005837)           30

# performances clean90 :     (5.3914141414141414, 4.488304093567253)            20

# performances clean80 :     (6.494949494949496, 6.128654970760238)             14



def prop_area (dataset,n):
    prop = []
    [gauss_B,gauss_M,area_B,area_M,liste_h] = courbes_param(dataset,n)
    for k in range (len(area_B)):
        gauss_B_k = gauss_B[k]
        gauss_M_k = gauss_M[k]
        area_B_k = area_B[k]
        area_M_k = area_M[k]
        h_k = liste_h[k]
        prop_k = []

        for i in range (1,n-1):
            gauss_B_k_i = gauss_B_k[:i+1]
            gauss_M_k_i = gauss_M_k[i+1:]
            area_B_k_i = simps(gauss_B_k_i , dx = h_k)
            area_M_k_i = simps(gauss_M_k_i , dx = h_k)
            prop_B_k_i = area_B_k_i/area_B_k
            prop_M_k_i = area_M_k_i/area_M_k
            prop_k.append((prop_B_k_i * prop_M_k_i)**1)
        prop.append(max(prop_k))
    return prop

def mise_en_forme_prop_area(dataset,n):
    liste_proportion = prop_area(dataset,n)
    pivot = max(liste_proportion)
    res = [ val/pivot  for val in liste_proportion]
    return res

def visualisation_importance_parametre(dataset):
    importance = mise_en_forme_prop_area(dataset,40)
    x = [ k for k in range (len(importance))]
    plt.bar(x,importance,color ='darkblue')
    plt.show()

def importance(dataset):
    return mise_en_forme_prop_area(dataset,40)

# objectif : coefficienter les paramètres pour une meilleure classification

# performances des données avec eval perf, n=200 sur la base wis :
#                                  mean perf normale   mean perf distance   nombre de paramètres

# non coefficienté,base normale :   (3.920454545454546, 3.321637426900584)      30
#     coefficienté,base normale :   (3.646464646464647, 3.190058479532162)      30

# non coefficienté clean90 :        (5.391414141414141, 4.488304093567253)      20
#     coefficienté clean90 :        (4.987373737373737, 4.116959064327485)      20

# non coefficienté clean80 :        (6.494949494949496, 6.128654970760238)      14
#     coefficienté clean80 :        (6.184343434343433, 5.356725146198833)      14

importances = importance(stand_base_wis)



# performances des données avec eval perf, n=200 sur la base wis :
#                        mean perf normale   mean perf distance

# importance +       :   (3.847222222222221 , 3.169590643274854 )
# importance *       :   (3.7525252525252553, 3.0906432748538015)
# immortance ** 0.75 :   (3.779040404040405 , 3.3216374269005855)
# immortance ** 1.5  :   (3.6250000000000004, 3.3011695906432754)
# importance **2     :   (3.5025252525252517, 3.333333333333334 )
